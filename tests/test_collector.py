"""Tests for umfrage.collector."""

from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from umfrage.collector import (
    CollectionSummary,
    GroupInfo,
    collect_all,
    collect_group,
    discover_questionnaire_groups,
    list_questionnaire_groups,
    resolve_config,
)
from umfrage.config_loader import ConfigError
from umfrage.generator import generate_questionnaire, write_metadata_file
from umfrage.models import (
    AnswerConfig,
    AnswerType,
    OrganizerInfo,
    Question,
    Questionnaire,
    RespondentField,
    Section,
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fill_response(
    xlsx_path: Path, name: str, institution: str, answers: dict
) -> None:
    """Write respondent info and answers into a copy of the base questionnaire."""
    wb = load_workbook(xlsx_path)
    ws = wb["Questionnaire"]
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        cell_text = str(val)
        if cell_text == "Name:":
            ws.cell(row=row, column=2).value = name
        elif cell_text == "Institution:":
            ws.cell(row=row, column=2).value = institution
        elif cell_text in answers:
            ws.cell(row=row, column=3).value = answers[cell_text]
    wb.save(xlsx_path)


SAMPLE_ANSWERS = {"G.Q1": 4, "G.Q2": "Yes", "G.Q3": "Great work!", "G.Q4": "Good", "T.Q1": 8}


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture()
def responses_folder(
    tmp_path: Path, sample_questionnaire: Questionnaire, sample_style
) -> Path:
    """Folder with 2 valid filled response files + metadata yaml."""
    folder = tmp_path / "responses"
    folder.mkdir()

    base_xlsx = tmp_path / "base.xlsx"
    generate_questionnaire(sample_questionnaire, sample_style, base_xlsx)
    write_metadata_file(
        sample_questionnaire,
        folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
    )

    for i, (name, inst) in enumerate([("Alice", "Univ A"), ("Bob", "Univ B")], start=1):
        resp_path = folder / f"response_{i}.xlsx"
        shutil.copy(base_xlsx, resp_path)
        _fill_response(resp_path, name, inst, SAMPLE_ANSWERS)

    return folder


# ── discover_questionnaire_groups ─────────────────────────────────────────────

class TestDiscoverGroups:
    def test_two_responses_form_one_group(self, responses_folder: Path) -> None:
        groups = discover_questionnaire_groups(responses_folder)
        assert len(groups) == 1
        paths = list(groups.values())[0]
        assert len(paths) == 2

    def test_result_files_are_skipped(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        # Place a results file in the folder; it must not be picked up
        results_file = responses_folder / "results_test_2024-01-01.xlsx"
        shutil.copy(list(responses_folder.glob("response_*.xlsx"))[0], results_file)
        groups = discover_questionnaire_groups(responses_folder)
        paths = list(groups.values())[0]
        assert results_file not in paths

    def test_corrupted_file_is_skipped(self, responses_folder: Path) -> None:
        bad = responses_folder / "bad.xlsx"
        bad.write_bytes(b"not an xlsx")
        groups = discover_questionnaire_groups(responses_folder)
        # Corrupted file should not appear in any group
        all_paths = [p for paths in groups.values() for p in paths]
        assert bad not in all_paths

    def test_empty_folder_returns_empty(self, tmp_path: Path) -> None:
        folder = tmp_path / "empty"
        folder.mkdir()
        assert discover_questionnaire_groups(folder) == {}


# ── resolve_config ────────────────────────────────────────────────────────────

class TestResolveConfig:
    def test_resolves_from_metadata_yaml(
        self, responses_folder: Path, sample_questionnaire: Questionnaire
    ) -> None:
        q = resolve_config(sample_questionnaire.config_hash(), responses_folder)
        assert q.title == sample_questionnaire.title
        assert q.config_hash() == sample_questionnaire.config_hash()

    def test_resolves_from_timestamped_metadata_yaml(
        self, tmp_path: Path, sample_questionnaire: Questionnaire, sample_style
    ) -> None:
        """CLI writes *_metadata_TIMESTAMP.yaml; resolve_config must find it."""
        folder = tmp_path / "ts"
        folder.mkdir()
        # Write the metadata file with a timestamp suffix, as the CLI does.
        ts_meta = folder / f"{sample_questionnaire.questionnaire_id()}_metadata_20260814_122339.yaml"
        write_metadata_file(sample_questionnaire, ts_meta)
        q = resolve_config(sample_questionnaire.config_hash(), folder)
        assert q.config_hash() == sample_questionnaire.config_hash()

    def test_config_override_takes_precedence(
        self, responses_folder: Path, sample_questionnaire: Questionnaire
    ) -> None:
        q = resolve_config("any_hash", responses_folder, config_override=sample_questionnaire)
        assert q.title == sample_questionnaire.title

    def test_unknown_hash_raises_config_error(self, responses_folder: Path) -> None:
        with pytest.raises(ConfigError, match="No questionnaire config"):
            resolve_config("deadbeef" * 8, responses_folder)

    def test_empty_folder_raises_config_error(self, tmp_path: Path) -> None:
        folder = tmp_path / "empty"
        folder.mkdir()
        with pytest.raises(ConfigError):
            resolve_config("any_hash", folder)


# ── collect_all ───────────────────────────────────────────────────────────────

class TestCollectAll:
    def test_produces_one_summary(
        self, responses_folder: Path, sample_style
    ) -> None:
        out_dir = responses_folder / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        assert len(summaries) == 1

    def test_summary_counts_correct(
        self, responses_folder: Path, sample_style
    ) -> None:
        out_dir = responses_folder / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        s = summaries[0]
        assert s.valid_count == 2
        assert s.skipped_count == 0
        assert s.total_files == 2

    def test_result_file_created(
        self, responses_folder: Path, sample_style
    ) -> None:
        out_dir = responses_folder / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        assert summaries[0].output_path is not None
        assert summaries[0].output_path.exists()

    def test_invalid_file_skipped_during_discovery(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        # A corrupted file is dropped by discover_questionnaire_groups (can't
        # read _meta), so it never enters any group. The two valid files should
        # still produce a successful result.
        bad = responses_folder / "bad.xlsx"
        bad.write_bytes(b"not xlsx")
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        # The bad file was silently dropped at discovery; valid files succeed.
        assert len(summaries) == 1
        assert summaries[0].valid_count == 2
        assert summaries[0].skipped_count == 0

    def test_two_questionnaires_produce_two_results(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        other_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        folder = tmp_path / "mixed"
        folder.mkdir()

        for q_obj, answers in [
            (sample_questionnaire, SAMPLE_ANSWERS),
            (other_questionnaire, {"F.Q1": 2}),
        ]:
            base = tmp_path / f"base_{q_obj.questionnaire_id()}.xlsx"
            generate_questionnaire(q_obj, sample_style, base)
            write_metadata_file(
                q_obj, folder / f"{q_obj.questionnaire_id()}_metadata.yaml"
            )
            resp = folder / f"resp_{q_obj.questionnaire_id()}.xlsx"
            shutil.copy(base, resp)
            _fill_response(resp, "Tester", "Org", answers)

        out_dir = tmp_path / "results"
        summaries = collect_all(folder, sample_style, out_dir)
        assert len(summaries) == 2

        result_files = sorted(out_dir.glob("results_*.xlsx"))
        assert len(result_files) == 2

    def test_output_dir_created_if_missing(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        out_dir = tmp_path / "new" / "dir"
        collect_all(responses_folder, sample_style, out_dir)
        assert out_dir.exists()


# ── Result workbook content ───────────────────────────────────────────────────

class TestResultWorkbook:
    def _run_and_open(self, responses_folder, sample_style, tmp_path):
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        wb = load_workbook(summaries[0].output_path, data_only=True)
        return wb["Results"]

    def test_results_sheet_exists(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        wb = load_workbook(summaries[0].output_path)
        assert "Results" in wb.sheetnames

    def test_institution_names_in_header_row(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        ws = self._run_and_open(responses_folder, sample_style, tmp_path)
        # Row 4 is the column header row (fixed cols: Section, Q-ID, Question, Scale/Comment)
        header_values = [ws.cell(row=4, column=c).value for c in range(1, 8)]
        assert "Univ A" in header_values or "Alice" in header_values
        assert "Univ B" in header_values or "Bob" in header_values

    def test_question_ids_in_result(
        self, responses_folder: Path, sample_style, tmp_path: Path,
        sample_questionnaire: Questionnaire
    ) -> None:
        ws = self._run_and_open(responses_folder, sample_style, tmp_path)
        col_b = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        for q in sample_questionnaire.all_questions():
            assert q.id in col_b, f"Question ID '{q.id}' not found in result column B"

    def test_title_in_row_one(
        self, responses_folder: Path, sample_style, tmp_path: Path,
        sample_questionnaire: Questionnaire
    ) -> None:
        ws = self._run_and_open(responses_folder, sample_style, tmp_path)
        title_cell = ws.cell(row=1, column=1).value
        assert sample_questionnaire.title in str(title_cell)

    def test_freetext_answer_cells_are_left_aligned(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        wb = load_workbook(summaries[0].output_path)
        ws = wb["Results"]
        # Find the G.Q3 row (FREETEXT question) by its Q-ID in column 2
        freetext_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=2).value == "G.Q3":
                freetext_row = r
                break
        assert freetext_row is not None, "G.Q3 (freetext) row not found in result sheet"
        # Respondent answer columns start at column 5
        for col in range(5, 7):
            cell = ws.cell(row=freetext_row, column=col)
            assert cell.alignment.horizontal == "left", (
                f"Expected left alignment for freetext answer col {col}, "
                f"got '{cell.alignment.horizontal}'"
            )

    def test_non_freetext_answer_cells_are_centered(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        wb = load_workbook(summaries[0].output_path)
        ws = wb["Results"]
        # G.Q1 is a SCALE question — its answer cells must remain centered
        scale_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=2).value == "G.Q1":
                scale_row = r
                break
        assert scale_row is not None, "G.Q1 (scale) row not found in result sheet"
        for col in range(5, 7):
            cell = ws.cell(row=scale_row, column=col)
            assert cell.alignment.horizontal == "center", (
                f"Expected center alignment for scale answer col {col}, "
                f"got '{cell.alignment.horizontal}'"
            )


class TestSourceFileRow:
    """Row 5 of the result sheet must list response filenames with hyperlinks."""

    def _result_ws(self, responses_folder, sample_style, tmp_path):
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        wb = load_workbook(summaries[0].output_path, data_only=True)
        return wb["Results"]

    def test_source_file_row_contains_filenames(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        ws = self._result_ws(responses_folder, sample_style, tmp_path)
        # Row 5 is the source-file row (row 4 = institution names)
        row5 = [ws.cell(row=5, column=c).value for c in range(1, 8)]
        assert any(
            str(v).endswith(".xlsx") for v in row5 if v is not None
        ), f"No .xlsx filename found in row 5: {row5}"

    def test_source_file_row_has_both_response_filenames(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        ws = self._result_ws(responses_folder, sample_style, tmp_path)
        row5 = [str(ws.cell(row=5, column=c).value or "") for c in range(1, 8)]
        assert any("response_1" in v for v in row5), "response_1.xlsx not found in row 5"
        assert any("response_2" in v for v in row5), "response_2.xlsx not found in row 5"

    def test_source_file_label_in_col_one(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        ws = self._result_ws(responses_folder, sample_style, tmp_path)
        label = ws.cell(row=5, column=1).value
        # Label must be non-empty (translated "Source file" / "Quelldatei")
        assert label is not None and str(label).strip() != ""

    def test_filename_cells_have_hyperlinks(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        """Filename cells in the source-file row must carry file:// hyperlinks."""
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        # Reload *without* data_only so hyperlink metadata is available
        wb = load_workbook(summaries[0].output_path)
        ws = wb["Results"]
        # Fixed cols: Section(1), Q-ID(2), Question(3), Scale(4) — inst cols start at 5
        hyperlinks_found = []
        for col in range(5, 5 + 2):  # 2 respondents
            cell = ws.cell(row=5, column=col)
            hyperlinks_found.append(cell.hyperlink is not None)
        assert all(hyperlinks_found), (
            f"Expected hyperlinks on all filename cells, got: {hyperlinks_found}"
        )

    def test_filename_cells_are_left_aligned(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        wb = load_workbook(summaries[0].output_path)
        ws = wb["Results"]
        for col in range(5, 5 + 2):
            cell = ws.cell(row=5, column=col)
            assert cell.alignment.horizontal == "left", (
                f"Expected left alignment on filename cell col {col}, "
                f"got '{cell.alignment.horizontal}'"
            )

    def test_hyperlinks_point_to_xlsx_files(
        self, responses_folder: Path, sample_style, tmp_path: Path
    ) -> None:
        out_dir = tmp_path / "out"
        summaries = collect_all(responses_folder, sample_style, out_dir)
        wb = load_workbook(summaries[0].output_path)
        ws = wb["Results"]
        for col in range(5, 5 + 2):
            cell = ws.cell(row=5, column=col)
            assert cell.hyperlink is not None
            target = str(cell.hyperlink.target)
            assert target.endswith(".xlsx"), (
                f"Hyperlink target does not end with .xlsx: {target}"
            )
            # Must be relative — no absolute path or file:// URI leaked
            assert not target.startswith("file://"), (
                f"Hyperlink should be relative, not a file:// URI: {target}"
            )
            assert not target.startswith("/"), (
                f"Hyperlink should be relative, not absolute: {target}"
            )

    def test_german_source_file_label_translated(
        self, tmp_path: Path, sample_questionnaire, sample_style
    ) -> None:
        """German questionnaire must show 'Quelldatei' as the row label."""
        import shutil
        from umfrage.generator import generate_questionnaire, write_metadata_file

        q = sample_questionnaire.model_copy(update={"language": "de"})
        folder = tmp_path / "de_responses"
        folder.mkdir()
        base = tmp_path / "base_de.xlsx"
        generate_questionnaire(q, sample_style, base)
        write_metadata_file(q, folder / f"{q.questionnaire_id()}_metadata.yaml")

        resp = folder / "antwort_1.xlsx"
        shutil.copy(base, resp)
        # Fill required fields so validation passes
        wb_resp = load_workbook(resp)
        ws_resp = wb_resp["Questionnaire"]
        for row in range(1, ws_resp.max_row + 1):
            v = ws_resp.cell(row=row, column=1).value
            if v == "Name:":
                ws_resp.cell(row=row, column=2).value = "Hans"
            elif v == "Institution:":
                ws_resp.cell(row=row, column=2).value = "Uni Berlin"
            elif v == "G.Q1":
                ws_resp.cell(row=row, column=3).value = 3
            elif v == "G.Q2":
                ws_resp.cell(row=row, column=3).value = "Ja"
            elif v == "G.Q4":
                ws_resp.cell(row=row, column=3).value = "Good"
            elif v == "T.Q1":
                ws_resp.cell(row=row, column=3).value = 5
        wb_resp.save(resp)

        out_dir = tmp_path / "out_de"
        summaries = collect_all(folder, sample_style, out_dir)
        assert summaries, "No collection summaries returned"
        wb_result = load_workbook(summaries[0].output_path, data_only=True)
        ws_result = wb_result["Results"]
        label = str(ws_result.cell(row=5, column=1).value or "")
        assert "Quelldatei" in label, (
            f"Expected 'Quelldatei' for German label, got: '{label}'"
        )


# ── on_invalid callback / force-include behaviour ─────────────────────────────

class TestForceInclude:
    """Tests for the on_invalid callback and force-inclusion of invalid files."""

    def _make_invalid_response(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        sample_style,
        *,
        omit_answer: str = "G.Q1",
    ) -> Path:
        """Return a response file with one required answer left blank."""
        base = tmp_path / "base.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, base)
        resp = tmp_path / "invalid_resp.xlsx"
        shutil.copy(base, resp)
        # Fill all answers except the one to omit
        partial = {k: v for k, v in SAMPLE_ANSWERS.items() if k != omit_answer}
        _fill_response(resp, "Eve", "Org X", partial)
        return resp

    def test_no_callback_skips_invalid(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        """Without on_invalid callback, invalid files are silently skipped (legacy)."""
        folder = tmp_path / "r"
        folder.mkdir()
        write_metadata_file(
            sample_questionnaire,
            folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
        )
        resp = self._make_invalid_response(tmp_path, sample_questionnaire, sample_style)
        shutil.copy(resp, folder / "resp.xlsx")

        out_dir = tmp_path / "out"
        summaries = collect_all(folder, sample_style, out_dir, on_invalid=None)
        s = summaries[0]
        assert s.skipped_count == 1
        assert s.force_included_count == 0
        assert s.valid_count == 0

    def test_include_callback_force_includes(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        """Callback returning 'include' causes the file to be force-included."""
        folder = tmp_path / "r"
        folder.mkdir()
        write_metadata_file(
            sample_questionnaire,
            folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
        )
        resp = self._make_invalid_response(tmp_path, sample_questionnaire, sample_style)
        shutil.copy(resp, folder / "resp.xlsx")

        out_dir = tmp_path / "out"
        summaries = collect_all(
            folder, sample_style, out_dir, on_invalid=lambda p, e: "include"
        )
        s = summaries[0]
        assert s.force_included_count == 1
        assert s.skipped_count == 0
        assert len(s.force_included_files) == 1
        assert s.output_path is not None and s.output_path.exists()

    def test_skip_callback_skips_file(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        """Callback returning 'skip' discards the file (same as legacy)."""
        folder = tmp_path / "r"
        folder.mkdir()
        write_metadata_file(
            sample_questionnaire,
            folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
        )
        resp = self._make_invalid_response(tmp_path, sample_questionnaire, sample_style)
        shutil.copy(resp, folder / "resp.xlsx")

        out_dir = tmp_path / "out"
        summaries = collect_all(
            folder, sample_style, out_dir, on_invalid=lambda p, e: "skip"
        )
        s = summaries[0]
        assert s.skipped_count == 1
        assert s.force_included_count == 0

    def test_force_included_answers_marked(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        """Missing answers in force-included files are filled with the marker."""
        folder = tmp_path / "r"
        folder.mkdir()
        write_metadata_file(
            sample_questionnaire,
            folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
        )
        resp = self._make_invalid_response(
            tmp_path, sample_questionnaire, sample_style, omit_answer="G.Q1"
        )
        shutil.copy(resp, folder / "resp.xlsx")

        style = sample_style.model_copy(update={"missing_answer_marker": "XXXXX"})
        out_dir = tmp_path / "out"
        summaries = collect_all(
            folder, style, out_dir, on_invalid=lambda p, e: "include"
        )
        wb = load_workbook(summaries[0].output_path, data_only=True)
        ws = wb["Results"]
        # Find the G.Q1 row (column B = Q-ID) and read the respondent answer column
        marker_found = False
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=2).value == "G.Q1":
                # institution columns start at fixed_cols+1 = 5
                val = ws.cell(row=row, column=5).value
                if str(val) == "XXXXX":
                    marker_found = True
                break
        assert marker_found, "Expected XXXXX marker in result for missing G.Q1 answer"

    def test_custom_marker_written_to_result(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        """A custom missing_answer_marker from StyleConfig appears in the result."""
        folder = tmp_path / "r"
        folder.mkdir()
        write_metadata_file(
            sample_questionnaire,
            folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
        )
        resp = self._make_invalid_response(
            tmp_path, sample_questionnaire, sample_style, omit_answer="G.Q1"
        )
        shutil.copy(resp, folder / "resp.xlsx")

        from umfrage.models import StyleConfig
        style = StyleConfig(missing_answer_marker="N/A")
        out_dir = tmp_path / "out"
        summaries = collect_all(
            folder, style, out_dir, on_invalid=lambda p, e: "include"
        )
        wb = load_workbook(summaries[0].output_path, data_only=True)
        ws = wb["Results"]
        marker_found = False
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=2).value == "G.Q1":
                val = ws.cell(row=row, column=5).value
                if str(val) == "N/A":
                    marker_found = True
                break
        assert marker_found, "Expected 'N/A' custom marker in result"

    def test_all_decision_includes_remaining(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        """When the callback returns 'all' for the first file, subsequent invalid
        files are included without calling the callback again."""
        folder = tmp_path / "r"
        folder.mkdir()
        base = tmp_path / "base.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, base)
        write_metadata_file(
            sample_questionnaire,
            folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
        )
        # Create two invalid response files (both missing G.Q1)
        for i in range(1, 3):
            resp = folder / f"resp_{i}.xlsx"
            shutil.copy(base, resp)
            partial = {k: v for k, v in SAMPLE_ANSWERS.items() if k != "G.Q1"}
            _fill_response(resp, f"User{i}", f"Org{i}", partial)

        call_count = [0]

        def _callback(path: Path, errors: list) -> str:
            call_count[0] += 1
            return "all"  # first call: include all

        out_dir = tmp_path / "out"
        summaries = collect_all(folder, sample_style, out_dir, on_invalid=_callback)
        s = summaries[0]
        # Callback should be called once (first file); second file auto-included
        assert call_count[0] == 1
        assert s.force_included_count == 2

    def test_none_decision_skips_remaining(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        """When the callback returns 'none' for the first file, subsequent invalid
        files are skipped without calling the callback again."""
        folder = tmp_path / "r"
        folder.mkdir()
        base = tmp_path / "base.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, base)
        write_metadata_file(
            sample_questionnaire,
            folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
        )
        for i in range(1, 3):
            resp = folder / f"resp_{i}.xlsx"
            shutil.copy(base, resp)
            partial = {k: v for k, v in SAMPLE_ANSWERS.items() if k != "G.Q1"}
            _fill_response(resp, f"User{i}", f"Org{i}", partial)

        call_count = [0]

        def _callback(path: Path, errors: list) -> str:
            call_count[0] += 1
            return "none"  # first call: skip all

        out_dir = tmp_path / "out"
        summaries = collect_all(folder, sample_style, out_dir, on_invalid=_callback)
        s = summaries[0]
        assert call_count[0] == 1
        assert s.skipped_count == 2
        assert s.force_included_count == 0


# ── list_questionnaire_groups ─────────────────────────────────────────────────

class TestListQuestionnaireGroups:
    def test_single_group_returned(
        self, responses_folder: Path, sample_questionnaire: Questionnaire
    ) -> None:
        groups = list_questionnaire_groups(responses_folder)
        assert len(groups) == 1
        g = groups[0]
        assert isinstance(g, GroupInfo)
        assert g.questionnaire_id == sample_questionnaire.questionnaire_id()
        assert g.title == sample_questionnaire.title
        assert g.file_count == 2
        assert len(g.files) == 2
        assert not g.unresolvable

    def test_two_groups_returned(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        other_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        folder = tmp_path / "mixed"
        folder.mkdir()
        for q_obj, answers in [
            (sample_questionnaire, SAMPLE_ANSWERS),
            (other_questionnaire, {"F.Q1": 2}),
        ]:
            base = tmp_path / f"base_{q_obj.questionnaire_id()}.xlsx"
            generate_questionnaire(q_obj, sample_style, base)
            write_metadata_file(q_obj, folder / f"{q_obj.questionnaire_id()}_metadata.yaml")
            resp = folder / f"resp_{q_obj.questionnaire_id()}.xlsx"
            shutil.copy(base, resp)
            _fill_response(resp, "Tester", "Org", answers)

        groups = list_questionnaire_groups(folder)
        assert len(groups) == 2
        ids = {g.questionnaire_id for g in groups}
        assert sample_questionnaire.questionnaire_id() in ids
        assert other_questionnaire.questionnaire_id() in ids

    def test_unresolvable_group(
        self, tmp_path: Path, sample_questionnaire: Questionnaire, sample_style
    ) -> None:
        folder = tmp_path / "no_meta"
        folder.mkdir()
        base = tmp_path / "base.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, base)
        # Intentionally skip write_metadata_file so config cannot be resolved.
        resp = folder / "resp.xlsx"
        shutil.copy(base, resp)
        _fill_response(resp, "Anon", "Org", SAMPLE_ANSWERS)

        groups = list_questionnaire_groups(folder)
        assert len(groups) == 1
        assert groups[0].unresolvable is True
        assert groups[0].file_count == 1

    def test_empty_folder_returns_empty(self, tmp_path: Path) -> None:
        folder = tmp_path / "empty"
        folder.mkdir()
        assert list_questionnaire_groups(folder) == []

    def test_config_file_populated_when_in_meta(
        self, tmp_path: Path, sample_questionnaire: Questionnaire, sample_style
    ) -> None:
        folder = tmp_path / "r"
        folder.mkdir()
        base = tmp_path / "base.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, base, config_file="survey.yaml")
        write_metadata_file(
            sample_questionnaire,
            folder / f"{sample_questionnaire.questionnaire_id()}_metadata.yaml",
        )
        resp = folder / "resp.xlsx"
        shutil.copy(base, resp)

        groups = list_questionnaire_groups(folder)
        assert groups[0].config_file == "survey.yaml"

    def test_config_file_none_when_not_in_meta(self, responses_folder: Path) -> None:
        # responses_folder uses generate_questionnaire without config_file argument
        groups = list_questionnaire_groups(responses_folder)
        assert groups[0].config_file is None

    def test_config_override_resolves_without_yaml(
        self, tmp_path: Path, sample_questionnaire: Questionnaire, sample_style
    ) -> None:
        folder = tmp_path / "no_meta"
        folder.mkdir()
        base = tmp_path / "base.xlsx"
        generate_questionnaire(sample_questionnaire, sample_style, base)
        resp = folder / "resp.xlsx"
        shutil.copy(base, resp)

        groups = list_questionnaire_groups(folder, config_override=sample_questionnaire)
        assert len(groups) == 1
        assert not groups[0].unresolvable
        assert groups[0].title == sample_questionnaire.title


# ── survey_filter ─────────────────────────────────────────────────────────────

class TestSurveyFilter:
    """Tests for the survey_filter parameter of collect_all()."""

    def _mixed_folder(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        other_questionnaire: Questionnaire,
        sample_style,
    ) -> Path:
        folder = tmp_path / "mixed"
        folder.mkdir()
        for q_obj, answers in [
            (sample_questionnaire, SAMPLE_ANSWERS),
            (other_questionnaire, {"F.Q1": 2}),
        ]:
            base = tmp_path / f"base_{q_obj.questionnaire_id()}.xlsx"
            generate_questionnaire(q_obj, sample_style, base)
            write_metadata_file(q_obj, folder / f"{q_obj.questionnaire_id()}_metadata.yaml")
            resp = folder / f"resp_{q_obj.questionnaire_id()}.xlsx"
            shutil.copy(base, resp)
            _fill_response(resp, "Tester", "Org", answers)
        return folder

    def test_filter_by_slug_selects_one_group(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        other_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        folder = self._mixed_folder(tmp_path, sample_questionnaire, other_questionnaire, sample_style)
        out_dir = tmp_path / "out"
        summaries = collect_all(
            folder, sample_style, out_dir,
            survey_filter=[sample_questionnaire.questionnaire_id()],
        )
        assert len(summaries) == 1
        assert summaries[0].questionnaire_id == sample_questionnaire.questionnaire_id()

    def test_filter_by_hash_prefix_selects_one_group(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        other_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        folder = self._mixed_folder(tmp_path, sample_questionnaire, other_questionnaire, sample_style)
        # Use the first 12 chars of the hash — should match exactly one group.
        hash_prefix = sample_questionnaire.config_hash()[:12]
        out_dir = tmp_path / "out"
        summaries = collect_all(
            folder, sample_style, out_dir,
            survey_filter=[hash_prefix],
        )
        assert len(summaries) == 1
        assert summaries[0].questionnaire_id == sample_questionnaire.questionnaire_id()

    def test_filter_unknown_token_produces_no_summaries(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        other_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        folder = self._mixed_folder(tmp_path, sample_questionnaire, other_questionnaire, sample_style)
        out_dir = tmp_path / "out"
        summaries = collect_all(
            folder, sample_style, out_dir,
            survey_filter=["nonexistent-survey-slug"],
        )
        assert summaries == []

    def test_no_filter_processes_all_groups(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        other_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        folder = self._mixed_folder(tmp_path, sample_questionnaire, other_questionnaire, sample_style)
        out_dir = tmp_path / "out"
        summaries = collect_all(folder, sample_style, out_dir)
        assert len(summaries) == 2

    def test_slug_collision_raises_value_error(
        self,
        tmp_path: Path,
        sample_questionnaire: Questionnaire,
        sample_style,
    ) -> None:
        """Two surveys with identical titles produce the same slug; filter must error."""
        folder = tmp_path / "collision"
        folder.mkdir()

        # Clone sample_questionnaire with a different version (different hash, same slug).
        twin = sample_questionnaire.model_copy(update={"version": "2.0"})
        assert twin.questionnaire_id() == sample_questionnaire.questionnaire_id()
        assert twin.config_hash() != sample_questionnaire.config_hash()

        for q_obj in (sample_questionnaire, twin):
            base = tmp_path / f"base_{q_obj.config_hash()[:8]}.xlsx"
            generate_questionnaire(q_obj, sample_style, base)
            write_metadata_file(q_obj, folder / f"{q_obj.config_hash()[:8]}_metadata.yaml")
            resp = folder / f"resp_{q_obj.config_hash()[:8]}.xlsx"
            shutil.copy(base, resp)
            _fill_response(resp, "User", "Org", SAMPLE_ANSWERS)

        out_dir = tmp_path / "out"
        with pytest.raises(ValueError, match="ambiguous"):
            collect_all(
                folder, sample_style, out_dir,
                survey_filter=[sample_questionnaire.questionnaire_id()],
            )

